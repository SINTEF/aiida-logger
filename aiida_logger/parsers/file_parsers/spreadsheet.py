from __future__ import absolute_import

import numpy as np

from aiida.plugins import DataFactory
from aiida_logger.parsers.file_parsers.base import BaseFileParser


class SpreadsheetParser(BaseFileParser):
    """Parser class for parsing spreadsheet in Office Open XML format."""
    def __init__(self, *args, **kwargs):
        super(SpreadsheetParser, self).__init__(*args, **kwargs)
        self.binary = True

    def _parse(self, file_handle):
        """Parse the content of the spreadsheet file as NumPy arrays."""

        from openpyxl import load_workbook
        # Load the spreadsheet using openpyxl
        wb = load_workbook(file_handle)
        # Locate what data configuration and extract
        if self.parameters['evolution'] == 'time':
            result = self._parse_time(wb)
        else:
            raise NotImplementedError

        return result

    def _parse_time(self, wb):
        """Parse the content which is indexed by time."""

        # Fetch active sheet
        ws = wb.active

        # Load time as NumPy array using the parameters
        times_raw = [
            timestep[0].value for timestep in ws[self.parameters['time_range']]
        ]
        reference_time = times_raw[0]
        times = np.array([(timestep - reference_time).total_seconds()
                          for timestep in times_raw])

        # Load data as NumPy array using the parameter
        data = np.array([[i.value for i in j]
                         for j in ws[self.parameters['data_range']]])

        # Combine the time steps and data
        time_data = np.column_stack((times, data))

        # Load labels
        labels = [
            label.value for label in ws[self.parameters['label_range']][0]
        ]
        # Override with manual labels
        for key, item in self.parameters['manual_label'].items():
            labels.insert(key, item)
        # Check that we have labels for all data (only check first row/column)
        if time_data[0].shape[0] != len(labels):
            raise ValueError(
                'Some of the data is missing labels, please correct the supplied parameters.'
            )
        # Extract comments
        comments = [[str(i.value) for i in j]
                    for j in ws[self.parameters['comment_range']]]
        # Compose data and metadata nodes
        data = DataFactory('array')()
        data.set_array('content', time_data)
        metadata = DataFactory('dict')(dict={
            'start_time': reference_time.utcnow(),
            'comments': comments,
            'labels': labels
        })

        return {'data': data, 'metadata': metadata}
